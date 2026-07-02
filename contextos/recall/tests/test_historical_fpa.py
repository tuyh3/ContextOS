"""Test historical FPA xlsx parser.

Reusable module. POC uses it to extract gold; v1 will use it to index FPA into RAG.
"""
import pytest
from pathlib import Path

from openpyxl import Workbook

SAMPLE_XLSX = Path(__file__).parent.parent.parent.parent / "data" / "poc" / "samples" / "sample-01-baseline-fpa.xlsx"


def test_parse_xlsx_returns_fparecord_list():
    if not SAMPLE_XLSX.exists():
        pytest.skip("No sample xlsx; user must place one first")
    from contextos.recall.historical_fpa import parse_fpa_xlsx, FPARecord
    records = parse_fpa_xlsx(SAMPLE_XLSX)
    assert len(records) > 0
    r = records[0]
    assert isinstance(r, FPARecord)
    assert r.function_name           # 接口/功能名(EI/EO/EQ/ILF/EIF 行)
    assert r.function_type in {"EI", "EO", "EQ", "ILF", "EIF", "OTHER"}


def test_extracts_acronyms_separately_from_tables(tmp_path):
    """Audit 2026-05-28 P1 fix: GUI/CNIC/BVS/SIM in function_name MUST
    end up in `referenced_acronyms`, NOT `referenced_tables`. Real samples
    from a large customer project don't reliably exercise strict SQL_TABLE
    extraction (FTR cells often contain numeric counts, not table names),
    so we use a synthetic xlsx here to verify the SEPARATION of the two paths.
    """
    p = tmp_path / "synthetic-fpa.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        wb.create_sheet("FPA")
        ws = wb.active
    assert ws is not None
    # Single-row header matching the canonical aliases (no split-header).
    ws.append(["功能名称", "功能类型", "FTR"])
    # Row 1: function_name has bare acronyms (GUI, BVS) — should land in
    # referenced_acronyms. FTR has a strict-format table (CB_CUST_INFO)
    # — should land in referenced_tables.
    ws.append(["Modify GUI for BVS provisioning", "EI", "CB_CUST_INFO"])
    # Row 2: function_name has only acronyms; FTR is empty.
    ws.append(["Validate CNIC via API", "EQ", ""])
    # Row 3: FTR has multiple strict tables, function_name has none.
    ws.append(["Settle balances", "EI", "CHNL_ORDER_HEADER, CB_BILL_LINE"])
    wb.save(str(p))

    from contextos.recall.historical_fpa import parse_fpa_xlsx
    records = parse_fpa_xlsx(p, requirement_id="synthetic-01")
    assert len(records) == 3

    # Row 1: GUI + BVS as acronyms; CB_CUST_INFO as table.
    assert "GUI" in records[0].referenced_acronyms
    assert "BVS" in records[0].referenced_acronyms
    assert "GUI" not in records[0].referenced_tables  # NOT a table
    assert "CB_CUST_INFO" in records[0].referenced_tables

    # Row 2: CNIC + API as acronyms; zero tables.
    assert "CNIC" in records[1].referenced_acronyms
    assert "API" in records[1].referenced_acronyms
    assert records[1].referenced_tables == []

    # Row 3: zero acronyms; both compound tables present.
    assert records[2].referenced_acronyms == []
    assert "CHNL_ORDER_HEADER" in records[2].referenced_tables
    assert "CB_BILL_LINE" in records[2].referenced_tables


def test_extracts_strict_tables_from_function_name(tmp_path):
    """Audit 2026-05-28 B refinement: strict tables can appear in
    function_name (not just FTR column). Sample-04 (FPA-Splitting UD
    FreeRG) embeds `SERVICE_CODE`, `LOAD_TYPE`, `RESERVE_FIELDS` etc.
    as backtick-quoted code snippets inside function_name; FTR cell only
    has a numeric count. Parser MUST find these by scanning both cells
    under the same strict rule.
    """
    p = tmp_path / "name-tables.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        wb.create_sheet("FPA")
        ws = wb.active
    assert ws is not None
    ws.append(["功能名称", "功能类型", "FTR"])
    # function_name has strict table; FTR has only a count (typical customer
    # sample-04 pattern).
    ws.append(["Add check for `SERVICE_CODE` in `filter.lua`", "ILF", 1])
    # function_name has acronyms only; FTR has strict table.
    ws.append(["Modify GUI for BVS validation", "EI", "CB_CUST_INFO"])
    # function_name has BOTH strict-table-shape AND acronym; FTR has both
    # (should dedupe via the not-in check).
    ws.append([
        "Update `LOAD_TYPE` rule for GUI panel", "EI",
        "LOAD_TYPE, CHNL_ORDER",
    ])
    wb.save(str(p))

    from contextos.recall.historical_fpa import parse_fpa_xlsx
    records = parse_fpa_xlsx(p, requirement_id="synthetic-name-tables")
    assert len(records) == 3

    # Row 1: SERVICE_CODE from function_name, no acronyms.
    assert "SERVICE_CODE" in records[0].referenced_tables
    assert records[0].referenced_acronyms == []

    # Row 2: GUI + BVS as acronyms; CB_CUST_INFO from FTR.
    assert "GUI" in records[1].referenced_acronyms
    assert "BVS" in records[1].referenced_acronyms
    assert "CB_CUST_INFO" in records[1].referenced_tables

    # Row 3: LOAD_TYPE + CHNL_ORDER (deduped — LOAD_TYPE in both name + FTR).
    assert "LOAD_TYPE" in records[2].referenced_tables
    assert "CHNL_ORDER" in records[2].referenced_tables
    assert records[2].referenced_tables.count("LOAD_TYPE") == 1, \
        "LOAD_TYPE in both function_name + FTR should be deduped"
    assert "GUI" in records[2].referenced_acronyms


def test_real_sample_has_extractable_records():
    """Real xlsx from a large customer project may yield zero strict-tables
    (FTR cells often hold numeric counts, not table names). What we DO require: the parser
    extracts >0 records, each with a function_name. SQL_TABLE evidence
    is a bonus when FTR data permits it."""
    if not SAMPLE_XLSX.exists():
        pytest.skip("No sample xlsx; run poc_t1_auto_select_samples.py first")
    from contextos.recall.historical_fpa import parse_fpa_xlsx
    records = parse_fpa_xlsx(SAMPLE_XLSX)
    assert len(records) > 0
    # Most records should at least have function_name. Acronyms / tables
    # are best-effort (data-dependent).
    with_names = [r for r in records if r.function_name]
    assert len(with_names) > 0


def test_fparecord_to_evidence_items():
    """FPARecord → EvidenceItem 转换,POC 用这个产 gold.json.

    Audit 2026-05-28 P1: INTERFACE items now carry `metadata.acronyms`;
    SQL_TABLE items only generated from STRICT-matched tables in
    `referenced_tables` (no longer from function_name acronyms).
    """
    if not SAMPLE_XLSX.exists():
        pytest.skip("No sample xlsx")
    from contextos.recall.historical_fpa import parse_fpa_xlsx
    from contextos.recall._poc_schema import EvidenceItem
    records = parse_fpa_xlsx(SAMPLE_XLSX)
    all_evidence = []
    for r in records:
        all_evidence.extend(r.to_evidence_items())
    assert len(all_evidence) > 0
    assert all(isinstance(e, EvidenceItem) for e in all_evidence)
    kinds = {e.kind for e in all_evidence}
    assert kinds.issubset({"INTERFACE", "SQL_TABLE", "SQL_COLUMN", "OTHER"})
    # Every record produces at least one INTERFACE item (the function itself).
    interface_items = [e for e in all_evidence if e.kind == "INTERFACE"]
    assert len(interface_items) > 0
    # INTERFACE.metadata must carry the acronyms field (may be empty list).
    for item in interface_items:
        assert "acronyms" in item.metadata
        assert isinstance(item.metadata["acronyms"], list)
