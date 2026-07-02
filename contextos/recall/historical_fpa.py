"""Parse historical FPA xlsx files.

POC use: extract gold-standard Impact Map evidence (Task 6).
v1 use:  index into RAG for "similar historical requirement" retrieval (Task ?).

The same FPARecord dataclass is the source-of-truth for both purposes. v1 will hold
out test-set requirements from the RAG index to prevent data leakage (see plan
§"评测方法论 + 数据泄漏防范").

XLSX structure expected (a large real customer project convention; adjust if other clients differ):
- Sheet "Function Point Analysis" or similar
- Columns: 功能编号 / 功能名称 / 功能类型(EI/EO/EQ/ILF/EIF) / 复杂度 / DET / FTR / RET / UFP / 备注
- 引用文件/表通常在 "FTR" 列或 "引用文件" 列,逗号分隔
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook  # pip install openpyxl


@dataclass
class FPARecord:
    """One row of an FPA Function Points sheet."""
    requirement_id: str          # 来自文件名 (sample-01) 或 xlsx 头部
    function_name: str           # 接口/功能名(原 xlsx 列)
    function_type: str           # EI / EO / EQ / ILF / EIF / OTHER
    complexity: str = ""         # 低 / 中 / 高
    det_count: int = 0           # Data Element Types
    ftr_count: int = 0           # File Types Referenced
    # referenced_tables: STRICT table candidates from FTR column only,
    # filtered by known prefix or underscored-compound shape (see
    # _extract_tables_strict). Used for SQL_TABLE evidence — must NOT
    # contain domain acronyms (GUI, CNIC, BVS, ...). Audit 2026-05-28
    # P1: bare ALL-CAPS tokens were polluting SQL_TABLE; now isolated to
    # `referenced_acronyms` below.
    referenced_tables: list[str] = field(default_factory=list)
    # referenced_acronyms: ALL-CAPS tokens from function_name. These are
    # interface-domain hints (telecom/programming acronyms) and become
    # metadata on the INTERFACE evidence — NOT separate SQL_TABLE items.
    referenced_acronyms: list[str] = field(default_factory=list)
    referenced_files: list[str] = field(default_factory=list)
    notes: str = ""              # 备注列原文
    raw_row: dict = field(default_factory=dict)  # 原始行所有字段(调试用)

    def to_evidence_items(self) -> list:
        """Convert one FPA record to one or more EvidenceItem objects.

        Granularity: xlsx gives INTERFACE-level + (strict) TABLE-level.
        INTERFACE 命中后 Task 10 用 fuzzy match。SQL_TABLE 只在有
        strict-match 的表名时才生成 — 不再把 function_name 里的 GUI/CNIC/
        BVS/API 当表(2026-05-28 audit P1 fix)。
        """
        # POC 口径 gold-evidence dataclass(非 v1 pydantic 契约)。Plan 01 把
        # impact_map.schema 重写成 v1 三维 pydantic 后,POC 构造迁到此冻结 shim。
        # Plan 09 评测重建时改产 v1 EvidenceItemWithDimensions(见 _poc_schema 文档)。
        from contextos.recall._poc_schema import EvidenceItem
        items = []
        # 1. 接口/功能本身 -> INTERFACE evidence (acronyms 在 metadata)
        if self.function_name:
            items.append(EvidenceItem(
                id=f"fpa-{self.function_type.lower()}-{re.sub(r'[^a-zA-Z0-9_]', '-', self.function_name)[:40]}",
                change_type="MODIFY" if "改" in self.notes or "modify" in self.notes.lower() else "NEW",
                kind="INTERFACE",
                target=self.function_name,
                file="(from-fpa-xlsx)",
                line_start=0,
                line_end=0,
                confidence="HIGH",
                evidence_source="historical-fpa",
                rationale=f"FPA xlsx 行: type={self.function_type} complexity={self.complexity} det={self.det_count} ftr={self.ftr_count}",
                metadata={
                    "raw": self.raw_row,
                    # acronyms = domain hints from function_name (e.g. GUI,
                    # CNIC, BVS, API). NOT promoted to standalone evidence
                    # because they are not literal SQL tables. Downstream
                    # consumers (Task 10) MAY use them as fuzzy hints.
                    "acronyms": self.referenced_acronyms,
                },
            ))
        # 2. 每个 STRICT-matched 表 -> SQL_TABLE evidence (audit P1: no longer
        # generates SQL_TABLE for bare function-name acronyms)
        for tbl in self.referenced_tables:
            items.append(EvidenceItem(
                id=f"fpa-table-{tbl}-from-{self.function_type.lower()}-{re.sub(r'[^a-zA-Z0-9_]', '-', self.function_name)[:30]}",
                change_type="DB",
                kind="SQL_TABLE",
                target=tbl,
                file="(from-fpa-xlsx)",
                line_start=0,
                line_end=0,
                confidence="HIGH",
                evidence_source="historical-fpa",
                rationale=f"FTR of {self.function_name}",
            ))
        return items


# ---------- xlsx 解析逻辑 ----------

# Note (2026-05-27 Task 6 deviation from plan): real customer FPA xlsx use a
# 2-row split header (row N = parent group labels, row N+1 = sub-labels under
# "Original Assessment"). Aliases below cover both the Chinese plan spec and
# the actual English/Chinese labels observed in samples (incl. typo
# "Detail Funcitons" preserved verbatim from customer templates).
_COLUMN_ALIASES = {
    "function_name": ["功能名称", "功能", "function name", "name", "接口名",
                      "detail funcitons", "detail functions", "子功能描述",
                      "子功能", "功能描述"],
    "function_type": ["功能类型", "类型", "function type", "type", "func type"],
    "complexity":    ["复杂度", "complexity"],
    "det":           ["det", "数据元素", "det 数"],
    "ftr":           ["ftr", "引用文件", "引用表", "ftr 数", "ftr/ret"],
    "ret":           ["ret", "记录元素"],
    "notes":         ["备注", "说明", "notes", "remark", "remarks"],
    "dev_type":      ["dev type", "开发类型", "新增/完善"],
    "function_module": ["function module", "功能模块"],
}

# Sheet titles known to be reference / summary tables, not FPA data rows.
_SKIP_SHEET_TITLES = {
    "Filling Instructions", "汇总", "Summary", "Pivot", "Cover", "封面",
}

_TYPE_NORMALIZE = {
    "ei": "EI", "eo": "EO", "eq": "EQ", "ilf": "ILF", "eif": "EIF",
    "外部输入": "EI", "外部输出": "EO", "外部查询": "EQ",
    "内部逻辑文件": "ILF", "外部接口文件": "EIF",
}


def _scan_row_for_aliases(row_vals: list[str], existing: dict[str, int] | None = None) -> dict[str, int]:
    """Scan one row for column aliases, return canonical -> col-idx mapping.

    `existing`: don't overwrite canonical keys already mapped (the outer
    multi-row scan prefers the FIRST row that defines a key — useful when
    parent + sub-header rows both contain related-but-different labels).
    """
    out: dict[str, int] = dict(existing or {})
    for canonical, aliases in _COLUMN_ALIASES.items():
        if canonical in out:
            continue
        lc_aliases = {a.lower() for a in aliases}
        for i, val in enumerate(row_vals):
            if val.strip().lower() in lc_aliases:
                out[canonical] = i
                break
    return out


def _find_header_row(sheet) -> tuple[int, dict[str, int]] | None:
    """寻找含 'function_name' + 'function_type' 列名的行 (允许 2 行 split header)。

    Returns (data_start_row_idx, alias_map). data_start_row_idx is the index
    where actual data begins (one past the last header row consumed).
    """
    max_scan = min(11, sheet.max_row + 1)
    for row_idx in range(1, max_scan):
        row_vals = [str(c.value or "").strip() for c in sheet[row_idx]]
        alias_map = _scan_row_for_aliases(row_vals)
        # Try merging with the NEXT row (split-header support). Only use the
        # 2-row form if the next row ACTUALLY contributes new alias keys —
        # otherwise we'd consume a data row as a sub-header. (Audit
        # 2026-05-28: synthetic test_extracts_acronyms_separately... showed
        # the first data row being dropped when the unconditional merge
        # produced len(merged) == len(alias_map).)
        if row_idx + 1 < max_scan:
            next_vals = [str(c.value or "").strip() for c in sheet[row_idx + 1]]
            merged = _scan_row_for_aliases(next_vals, existing=alias_map)
            if (
                len(merged) > len(alias_map)
                and "function_name" in merged
                and "function_type" in merged
            ):
                return row_idx + 2, merged  # data begins one past sub-header
        # Single-row header fallback
        if "function_name" in alias_map and "function_type" in alias_map:
            return row_idx + 1, alias_map
    return None


def _extract_tables_from_ftr(ftr_text: str) -> list[str]:
    """从 FTR / function_name 文本抽出表名 / API 名,容忍各种分隔符 / 前后缀。

    Strict rule (audit 2026-05-28 P1 fix): a token qualifies as a table
    candidate only if it has multi-segment structure (contains '_') OR
    starts with a KNOWN customer table prefix. Bare ALL-CAPS acronyms
    (GUI, CNIC, BVS, API, SIM, GMLC, OCS, ...) are NOT tables; they go
    through `_extract_acronyms_from_name` and end up in INTERFACE
    metadata instead.

    Examples (matched as table):  CMPK_CMPT_USER, DQ_HQ_HNAN_PAYMONEY,
                                  CB_CUST_INFO, CHNL_ORDER_HEADER
    Examples (NOT matched):       GUI, CNIC, BVS, API, SIM
    """
    if not ftr_text:
        return []
    raw = re.findall(r"\b[A-Z][A-Z0-9_]{2,}\b", str(ftr_text))
    seen: list[str] = []
    for t in raw:
        if t in seen:
            continue
        if "_" in t or any(t.startswith(p) for p in _KNOWN_TABLE_PREFIXES):
            seen.append(t)
    return seen


# Known customer SQL table prefixes (from infer_module_scope's
# table_prefix_to_module mapping). Tokens starting with these qualify as
# SQL_TABLE candidates even without an underscore (rare in practice).
_KNOWN_TABLE_PREFIXES = (
    "CB_", "CHNL_", "ORD_", "PORTAL_", "SEC_", "SSO_", "UPC_", "IRSC_",
    "CMPK_", "DQ_", "PA_", "DC_", "TF_",
)

# Domain acronyms commonly seen in customer function_names. These are NOT
# tables — they're telecom/programming concepts that the strict extraction
# correctly excludes from SQL_TABLE. Listed for documentation only; the
# actual filter is structural (must have '_' or known prefix).
_KNOWN_DOMAIN_ACRONYMS = frozenset({
    "GUI", "CNIC", "BVS", "API", "SIM", "GMLC", "OCS", "USSD", "SMS", "MMS",
    "IVR", "SQL", "XML", "JSON", "REST", "HTTP", "HTTPS", "CSV", "PDF",
    "VOLTE", "FTTH", "CPN", "BYN", "MZA", "MBVS", "ARPU",
})


def _extract_acronyms_from_name(function_name: str) -> list[str]:
    """ALL-CAPS short tokens from function_name. End up in INTERFACE
    metadata, not standalone SQL_TABLE evidence. Audit 2026-05-28 P1.
    """
    if not function_name:
        return []
    raw = re.findall(r"\b[A-Z][A-Z0-9_]{1,}\b", str(function_name))
    seen: list[str] = []
    for t in raw:
        # Skip if it's structurally table-like (those go to referenced_tables).
        if "_" in t or any(t.startswith(p) for p in _KNOWN_TABLE_PREFIXES):
            continue
        if t in seen:
            continue
        seen.append(t)
    return seen


def parse_fpa_xlsx(xlsx_path: Path, requirement_id: str | None = None) -> list[FPARecord]:
    """Parse an FPA xlsx, return list of FPARecord (one per data row).

    requirement_id default: derived from filename (sample-NN-baseline-fpa.xlsx -> sample-NN).
    """
    if requirement_id is None:
        name = xlsx_path.stem
        # sample-01-baseline-fpa -> sample-01
        m = re.match(r"(sample-\d+)", name)
        requirement_id = m.group(1) if m else name
    # Narrow Optional[str] → str for pyright: by here we've either kept the
    # caller-provided requirement_id or derived one from the filename. Both
    # paths produce str.
    assert requirement_id is not None

    wb = load_workbook(xlsx_path, data_only=True)
    try:
        records: list[FPARecord] = []
        for sheet in wb.worksheets:
            if sheet.title in _SKIP_SHEET_TITLES:
                continue
            header = _find_header_row(sheet)
            if header is None:
                continue
            data_start_row, cols = header
            for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                name_idx = cols["function_name"]
                if name_idx >= len(row):
                    continue
                name = str(row[name_idx] or "").strip()
                if not name or name.lower() in {"小计", "合计", "total", "subtotal"}:
                    continue
                ftype_raw = str(row[cols["function_type"]] or "").strip().lower() if cols.get("function_type", -1) < len(row) else ""
                ftype = _TYPE_NORMALIZE.get(ftype_raw, "OTHER")
                ftr_cell = row[cols["ftr"]] if "ftr" in cols and cols["ftr"] < len(row) else None
                # Audit 2026-05-28 P1 + B refinement: SEPARATE table extraction
                # from acronym extraction by STRUCTURAL RULE, not by source
                # location.
                # - Tables (strict: has '_' or known prefix) can appear in
                #   EITHER FTR cell OR function_name. Customer xlsx vary: some
                #   put table names in FTR; others (e.g. sample-04 Splitting
                #   UD FreeRG) embed them in function_name as backtick-quoted
                #   code snippets (`SERVICE_CODE`, `LOAD_TYPE`,
                #   `RESERVE_FIELDS`). Scanning both prevents Task 10 from
                #   getting zero SQL_TABLE evidence on the latter style.
                # - Acronyms (bare ALL-CAPS, no '_', not a known prefix) come
                #   only from function_name and go to INTERFACE.metadata —
                #   never standalone SQL_TABLE items.
                ftr_tables = _extract_tables_from_ftr(str(ftr_cell or ""))
                name_tables = _extract_tables_from_ftr(str(name))
                tables: list[str] = []
                for t in ftr_tables + name_tables:
                    if t not in tables:
                        tables.append(t)
                acronyms = _extract_acronyms_from_name(name)
                # Assign-then-narrow pattern: openpyxl cell values come back as
                # a wide union (float | Decimal | str | CellRichText | datetime
                # | ArrayFormula | ...). Pyright can't narrow subscript
                # expressions through isinstance — but it CAN narrow a named
                # variable. So we bind to det_val / ftr_val first.
                det_idx = cols.get("det", -1)
                det_val = row[det_idx] if 0 <= det_idx < len(row) else None
                try:
                    det_count = int(det_val) if isinstance(det_val, (int, float)) else 0
                except (TypeError, ValueError):
                    det_count = 0

                ftr_idx = cols.get("ftr", -1)
                ftr_val = row[ftr_idx] if 0 <= ftr_idx < len(row) else None
                try:
                    ftr_count = int(ftr_val) if isinstance(ftr_val, (int, float)) else len(tables)
                except (TypeError, ValueError):
                    ftr_count = len(tables)
                notes_idx = cols.get("notes", -1)
                notes = str(row[notes_idx] or "").strip() if 0 <= notes_idx < len(row) else ""
                comp_idx = cols.get("complexity", -1)
                complexity = str(row[comp_idx] or "").strip() if 0 <= comp_idx < len(row) else ""
                records.append(FPARecord(
                    requirement_id=requirement_id,
                    function_name=name,
                    function_type=ftype,
                    complexity=complexity,
                    det_count=det_count,
                    ftr_count=ftr_count,
                    referenced_tables=tables,
                    referenced_acronyms=acronyms,
                    notes=notes,
                    raw_row={k: row[v] for k, v in cols.items() if v < len(row)},
                ))
        return records
    finally:
        wb.close()


# ---------- requirement_summary 提取 (docx 优先, xlsx 兜底) ----------

DESCRIPTION_SHEET_NAMES = {"Description", "Background", "概述", "需求描述",
                          "需求", "Requirement", "Overview"}


def extract_summary_from_xlsx(xlsx_path: Path) -> str:
    """Pull requirement_summary text from xlsx description sheet or first long cells.
    Used when no docx pair available (DEGRADED sample).
    """
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        for sheet in wb.worksheets:
            if sheet.title in DESCRIPTION_SHEET_NAMES:
                txts = []
                for row in sheet.iter_rows(max_row=20, values_only=True):
                    for v in row:
                        s = str(v or "").strip()
                        if len(s) > 15 and not s.lower().startswith(("date:", "version", "author")):
                            txts.append(s)
                if txts:
                    return "\n\n".join(txts[:5])
        sheet = wb.worksheets[0]
        txts = []
        for row in sheet.iter_rows(max_row=15, values_only=True):
            for v in row:
                s = str(v or "").strip()
                if len(s) > 30 and not s.lower().startswith(("date:", "version", "author", "function", "complexity")):
                    txts.append(s)
        return "\n\n".join(txts[:3]) if txts else re.sub(r"[-_]", " ", xlsx_path.stem)
    finally:
        wb.close()


def extract_summary_from_docx(docx_path: Path) -> str:
    """Pull requirement_summary from first ~10 non-empty paragraphs of .docx."""
    from docx import Document
    doc = Document(str(docx_path))
    paras = []
    for p in doc.paragraphs:
        text = p.text.strip()
        if not text or len(text) < 10:
            continue
        if re.match(r"^(version|date|author|approved by)\s*[:|]", text, re.IGNORECASE):
            continue
        paras.append(text)
        if len(paras) >= 10:
            break
    return "\n\n".join(paras) if paras else "(empty docx)"


def infer_module_scope(
    records: list[FPARecord], summary_text: str | None = None
) -> list[str]:
    """Heuristic: from referenced_tables + function_name + optional summary text,
    infer business module names.

    Looks for known customer module hints in table names (prefix conventions like
    CB_*=customer, CHNL_*=channel, ...), function_name lowercased substrings, and
    (optionally) the requirement_summary text.

    The records-derived modules can be any in `known`. The summary-text-derived
    modules are restricted to `corpus_modules` — modules that have an actual
    presence in the indexed corpus (demoproj/<module>/CLAUDE.md). This prevents the
    summary scan from inflating module_scope with generic business terms like
    "BOSS APIs" or "Admin portal" that don't map to indexable sources, which
    would artificially lower recall metrics.
    """
    known = {"cust", "channel", "irsc", "order", "portal", "sec", "sso", "upc",
             "billing", "boss", "vcms", "ftth", "volte", "ocs", "crm", "cpn"}
    # Subset that has actual customer source-tree CLAUDE.md presence (the indexed
    # corpus). Summary scans are restricted to this set.
    corpus_modules = {"acctmanm", "bin", "channel", "cust", "irsc", "order",
                      "sec", "upc"}
    table_prefix_to_module = {
        "CB_": "cust", "CHNL_": "channel", "ORD_": "order", "PORTAL_": "portal",
        "SEC_": "sec", "SSO_": "sso", "UPC_": "upc", "IRSC_": "irsc",
    }
    mods: set[str] = set()
    for r in records:
        for tbl in r.referenced_tables:
            for prefix, mod in table_prefix_to_module.items():
                if tbl.upper().startswith(prefix):
                    mods.add(mod)
        name_l = r.function_name.lower()
        for m in known:
            if m in name_l:
                mods.add(m)
    if summary_text:
        # Restrict text-scan candidates to modules that have actual corpus
        # presence. Avoids over-counting terms like "BOSS" or "portal" that
        # exist as business concepts but lack indexed source coverage.
        text_l = summary_text.lower()
        for m in corpus_modules:
            if m in text_l:
                mods.add(m)
    return sorted(mods)


# ---------- 给 v1 RAG 留的接口 ----------

def fpa_records_for_rag(xlsx_paths: list[Path], exclude_ids: set[str] | None = None) -> list[FPARecord]:
    """v1 RAG 用: 加载多个 xlsx 的 FPA,排除 hold-out 测试集。

    POC 阶段**不调用此函数**(POC 用 parse_fpa_xlsx 单文件 + 当 gold)。
    v1 阶段调用,exclude_ids = 当前测试 batch 的 requirement_id,防止数据泄漏。
    """
    exclude = exclude_ids or set()
    all_records = []
    for p in xlsx_paths:
        for r in parse_fpa_xlsx(p):
            if r.requirement_id not in exclude:
                all_records.append(r)
    return all_records
